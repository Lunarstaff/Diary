from selenium import webdriver
from selenium.webdriver.common.touch_actions import TouchActions
import time
import openpyxl

wx_rul = "https://testmy.orangebank.com.cn/WeiXin/773A50B48B2286170DE9A2D3FF014EAF92D09B8C815C0E0C51741AAA2FD67D" \
             "1B96C10BD04CA2737C73F42DC5CAE1AC6383D4217E902E8DC6C0A7A92A3767427EF47ADAB1B02F2BA828AACA271C56C0E10D12" \
             "D927DE2DDF5252C27C145BF42C38A810BD7CFDC2DAB28005E17A459EA7FF2363ADB5C01BEB85C7CE809B977013162D54FA5826" \
             "A647966ADDF84F3E99B6195E18E2CE916137A91B1C1820539D9D60354C504A15D997C3386D19641DD832DB95715AFF7EA0E8B0" \
             "D77C9625EF6AE9193D42D869233E79BFC031DE443B6B8E717C8CD9B5159F19DF8544AB7EE3F13FCF_bespeakService_1_1.do"

# 给浏览器定义启动参数
chrome_start_options = webdriver.ChromeOptions()
option_args = {"deviceName": "iPhone 6"}
chrome_start_options.add_experimental_option("mobileEmulation", option_args)
chrome_start_options.add_experimental_option("w3c", False)

# 带参启动浏览器
chrome_001 = webdriver.Chrome(chrome_options=chrome_start_options)
chrome_001.maximize_window()

# 从模板表中读取证件号录入
data_plate_file_f = "C:\\E-Data-File\\腾讯课堂\\Diary\\内容分类\\python-selenium\\微信预约开户数据模板.xlsx"
data_plate_file_xl = openpyxl.load_workbook(data_plate_file_f)
data_max_row = data_plate_file_xl.worksheets[0].max_row

# 根据数据模板进入主循环
for data_row in range(1, data_max_row):
    chrome_001.get(wx_rul)
    time.sleep(2)

    # 进入预约开户
    account_open_button = chrome_001.find_element_by_xpath('//*[@id="app"]/article/button[1]')
    account_open_button.click()
    time.sleep(1)
    alart_submit_button = chrome_001.find_element_by_xpath('//*[@id="alert_widow_remark"]/div[2]/a[2]')
    alart_submit_button.click()
    time.sleep(2)

    # 选择单位证件类型为 其他证件
    com_id_other_type = chrome_001.find_element_by_xpath('//*[@id="proveReturn"]')
    com_id_other_type.click()
    time.sleep(1)
    com_id_other_type_chose = chrome_001.find_element_by_xpath('//*[@id="proveTypeName"]')
    com_id_other_type_chose.click()
    time.sleep(1)
    # 定义触控，下拉框选择
    action_001 = TouchActions(chrome_001)
    com_id_other_type_00 = chrome_001.find_element_by_xpath('//*[@id="oneLevelContain"]/ul/li[1]')
    action_001.scroll_from_element(com_id_other_type_00, 0, 200)
    action_001.perform()
    time.sleep(2)
    # 单位其他证件类型选择OK
    com_id_other_type_ok = chrome_001.find_element_by_xpath('/html/body/div[9]/div/header/a[2]')
    com_id_other_type_ok.click()

    com_id_number = data_plate_file_xl.worksheets[0].cell(data_row+1, 2).value  # excel中行列是从1开始数
    com_id_number_input = chrome_001.find_element_by_xpath('//*[@id="proveNoNew"]')
    com_id_number_input.send_keys(com_id_number)

    try:
        # 是否法人亲临
        keyperson_value = data_plate_file_xl.worksheets[0].cell(data_row+1, 3).value
        if_keyperson_button = chrome_001.find_element_by_xpath('//*[@id="ifLegalSpan"]')
        if keyperson_value == "是":
            if_keyperson_button.click()
        elif keyperson_value == "否":
            pass
        # 法人证件类型选择其他
        keyperson_data_id_type = chrome_001.find_element_by_xpath('//*[@id="operatorCertTypeName"]')
        keyperson_data_id_type.click()
        time.sleep(1)
    except:
        # 如果弹出上次填单提示框
        if_cachedata_frame = chrome_001.find_element_by_xpath('/html/body/div[2]/div')
        cancell_button = chrome_001.find_element_by_xpath('/html/body/div[2]/div/div/a[1]')
        cancell_button.click()
        time.sleep(1)
        # 是否法人亲临
        keyperson_value = data_plate_file_xl.worksheets[0].cell(data_row+1, 3).value
        if_keyperson_button = chrome_001.find_element_by_xpath('//*[@id="ifLegalSpan"]')
        if keyperson_value == "是":
            if_keyperson_button.click()
            time.sleep(1)
        elif keyperson_value == "否":
            pass
        # 法人证件类型选择其他
        keyperson_data_id_type = chrome_001.find_element_by_xpath('//*[@id="operatorCertTypeName"]')
        keyperson_data_id_type.click()
        time.sleep(1)

    # 法人信息录入
    # 法人信息在excel表中的位置是 第2行开始 第4到第9列
    keyperson_data = {
        "法人证件类型": data_plate_file_xl.worksheets[0].cell(data_row+1, 4).value[0],
        "法人姓名": data_plate_file_xl.worksheets[0].cell(data_row+1, 5).value,
        "法人证件号码": data_plate_file_xl.worksheets[0].cell(data_row+1, 6).value,
        "法人证件到期日期": str(data_plate_file_xl.worksheets[0].cell(data_row+1, 7).value),
        "法人国籍": data_plate_file_xl.worksheets[0].cell(data_row+1, 8).value[0:3],
        "法人电话号码": data_plate_file_xl.worksheets[0].cell(data_row+1, 9).value
                      }

    # 定义触控，下拉框选择
    action_002 = TouchActions(chrome_001)
    keyperson_data_id_type_00 = chrome_001.find_element_by_xpath('//*[@id="oneLevelContain"]/ul/li[1]')
    action_002.scroll_from_element(keyperson_data_id_type_00, 0, 200)
    action_002.perform()
    time.sleep(1)
    keyperson_data_id_type_ok = chrome_001.find_element_by_xpath('/html/body/div[9]/div/header/a[2]')
    keyperson_data_id_type_ok.click()
    time.sleep(1)

    # 法人姓名录入
    keyperson_data_name = chrome_001.find_element_by_xpath('//*[@id="operatorName"]')
    keyperson_data_name.send_keys(keyperson_data["法人姓名"])
    time.sleep(1)

    # 法人证件号录入
    keyperson_data_id = chrome_001.find_element_by_xpath('//*[@id="operatorCertNo"]')
    keyperson_data_id.send_keys(keyperson_data["法人证件号码"])
    time.sleep(1)

    # 法人证件到期日期选择长期
    keyperson_data_id_date = chrome_001.find_element_by_xpath('//*[@id="operatorCertExpireSpan"]')
    keyperson_data_id_date.click()
    time.sleep(1)

    # 法人手机号录入
    keyperson_data_phone = chrome_001.find_element_by_xpath('//*[@id="operatorMobile"]')
    keyperson_data_phone.send_keys(keyperson_data["法人电话号码"])
    time.sleep(1)

    # 验证码录入
    image_code = chrome_001.find_element_by_xpath('//*[@id="TokenCode"]')
    image_code.send_keys("1111")
    time.sleep(1)

    # 短信验证码录入
    phone_code = chrome_001.find_element_by_xpath('//*[@id="operatorCheckCode"]')
    phone_code.send_keys("111111")
    time.sleep(1)

    # 下一页
    next_step_button = chrome_001.find_element_by_xpath('/html/body/div[1]/section[2]/a')
    next_step_button.click()
    time.sleep(2)













